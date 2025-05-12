from openpyxl import load_workbook

# Open an existing workbook
wb = load_work(r'C:\Users\Usyk\Desktop\python\excel.xlsx')

# Open the active worksheet
ws = wb.active

# Return a list of names of all worksheet
ws_names = wb.sheetnames

# Iterate through all the rows or columns of a file
ws_rows = ws.rows 
ws_columns = ws.columns

# Return the value of a cell
ws['B3'].value

# Create new worksheets
ws1 = wb.create_sheet('Mysheet')  # insert at the end (default)
# or
ws2 = wb.create_sheet('Mysheet', 0)  # insert at first position
# or 
ws3 = wb.create_sheet('Mysheet', -1)  # insert at penultimate position


# The simplest and safest way to save a workbook
wb.save('excel.xlsx')  # overwrite the `excel.xlse` file
# or 
wb.save('another.xlsx')  # save as an another file

from openpyxl import Workbook

# Create a file
wb = Workbook()

# Change the name of the sheet
ws.title = 'New title'

# Create copies of worksheets within a single workbool:
source = wb.active
target = wb.copy_worksheet(source)

# Add data to the last empty row
ws.append(['fist cell', 'second cell', 'third cell'])






