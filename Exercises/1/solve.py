import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter 

with open('data.json', 'r', encoding='utf-8') as file:
  data = json.load(file)

wb = Workbook()
ws = wb.active

ws.append(['Name', 'Age', 'Weight'])

for person in data:
  ws.append(list(person.values()))

for col in range(2,4):
  char = get_column_letter(col)
  ws[char+'5'].value = f'=AVERAGE({char+'2'}:{char+'4'})'

wb.save('new.xlsx')
