from openpyxl import Workbook
from openpyxl.styles.fonts import Font

wb = Workbook()
ws = wb.active 

ws['A1'].value = 'Name'
ws['A1'].font = Font(bold=True, color='00FF6600')

wb.save('new.xlsx')
