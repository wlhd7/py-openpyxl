from openpyxl import Workbook
from openpyxl.styles.fonts import Font

wb = Workbook()
ws = wb.active 

ws['A1'].value = 'Name'
ws['A1'].font = Font(bold=True)

wb.save('new.xlsx')
