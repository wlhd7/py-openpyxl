from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active

ws['A2'].fill = PatternFill(
  fgColor='FFFFFF00'
  start_color='FFFF00', 
  end_color='FFFF00',
  fill_type='solid')

if ws['A2'].fill.fgColor.rgb == 'FFFFFF00': return('fgColor')
if ws['A2'].fill.start_color == 'FFFFFF00': return('start_color')

wb.save('new.xlsx')

