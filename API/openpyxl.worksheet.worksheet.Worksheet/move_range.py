from openpyxl import Workbook 
from openpyxl.utils import get_column_letter

wb = Workbook() 
ws = wb.active 

for row in range(1,4):
  for col in range(1,4):
    ws[get_column_letter(col)+str(row)].value = 'x'

ws.move_range('A1:D4', rows=2, cols=3)

wb.save('new.xlsx')
