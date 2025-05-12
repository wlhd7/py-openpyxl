from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.merge_cells('A1:E1')
ws.unmerge_cells('A1:E1')
